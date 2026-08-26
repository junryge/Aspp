# -*- coding: utf-8 -*-
"""
요청/응답 관리 시스템 (FAB 개선요청 이력 관리)
--------------------------------------------------
- 단일 파일 Flask 앱 (폐쇄망 / pip-only 전제, Docker·Node 불필요)
- DB: SQLite 파일 하나 (reqlog.db)
- 외부 CDN 리소스 전혀 사용 안 함 (HTML/CSS/JS 내장)

실행:
    pip install flask openpyxl
    python app.py
    -> http://<서버IP>:10500

의존성: flask, openpyxl  (그 외 표준 라이브러리만 사용)
"""

import csv
import io
import json
import os
import re
import sqlite3
import sys
import uuid
from datetime import datetime, date, timedelta

from flask import Flask, g, jsonify, request, send_file, Response

# ----------------------------------------------------------------------------
# 설정
# ----------------------------------------------------------------------------
HOST = os.environ.get("REQLOG_HOST", "0.0.0.0")
PORT = int(os.environ.get("REQLOG_PORT", "10500"))
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.environ.get("REQLOG_DB", os.path.join(BASE_DIR, "reqlog.db"))
UPLOAD_DIR = os.environ.get("REQLOG_UPLOAD", os.path.join(BASE_DIR, "uploads"))

CATEGORIES = ["요청", "제안", "확인", "이슈"]
STATUSES = ["대기", "검토중", "적용완료", "보류", "반려"]

# 첨부 — 사진은 썸네일로 보이고, 문서는 파일칩으로 보인다.
# 어느 쪽이든 엑셀에는 파일 자체가 아니라 '건수 + 파일명'만 들어간다.
ALLOWED_IMG = {"png", "jpg", "jpeg", "gif", "bmp", "webp"}
ALLOWED_DOC = {
    "xlsx", "xls", "xlsm", "xlsb", "csv", "tsv",           # 엑셀/표
    "doc", "docx", "ppt", "pptx", "pdf",                    # 문서
    "hwp", "hwpx",                                          # 한글
    "txt", "log", "md", "json", "xml", "yaml", "yml", "sql", "ini", "conf",
    "zip", "7z", "tar", "gz", "tgz",                        # 압축
    "eml", "msg",                                           # 메일
}
ALLOWED_EXT = ALLOWED_IMG | ALLOWED_DOC

# 실행 가능한 확장자는 받지 않는다 (사내 배포물에 악성파일 얹히는 경로 차단)
BLOCKED_EXT = {"exe", "bat", "cmd", "com", "scr", "msi", "dll", "sh", "ps1",
               "vbs", "js", "jse", "wsf", "hta", "cpl", "reg", "jar", "lnk",
               "svg", "html", "htm"}

MIME_BY_EXT = {
    "png": "image/png", "jpg": "image/jpeg", "jpeg": "image/jpeg",
    "gif": "image/gif", "bmp": "image/bmp", "webp": "image/webp",
    "pdf": "application/pdf",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "pptx": "application/vnd.openxmlformats-officedocument.presentationml.presentation",
    "csv": "text/csv", "txt": "text/plain", "log": "text/plain",
    "zip": "application/zip",
}

MAX_MB = int(os.environ.get("REQLOG_MAXMB", "32"))

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = MAX_MB * 1024 * 1024  # 파일당 업로드 상한


# ----------------------------------------------------------------------------
# DB
# ----------------------------------------------------------------------------
SCHEMA = """
CREATE TABLE IF NOT EXISTS items (
    id           INTEGER PRIMARY KEY AUTOINCREMENT,
    seq          INTEGER,                    -- 화면에 보이는 No.
    category     TEXT NOT NULL DEFAULT '요청',
    requester    TEXT NOT NULL DEFAULT '',
    request_date TEXT,                       -- YYYY-MM-DD
    content      TEXT NOT NULL DEFAULT '',
    target       TEXT NOT NULL DEFAULT '',   -- 대상 시스템/FAB 태그 (콤마 구분)
    status       TEXT NOT NULL DEFAULT '대기',
    applied_date TEXT,
    confirmed_at TEXT,                      -- 고객 최종완료 확정 시각 (NULL이면 미확정)
    confirmed_by TEXT,                      -- 확정한 사람
    created_at   TEXT NOT NULL,
    updated_at   TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS responses (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    item_id       INTEGER NOT NULL,
    responder     TEXT NOT NULL DEFAULT '',
    response_date TEXT,
    content       TEXT NOT NULL DEFAULT '',
    created_at    TEXT NOT NULL,
    updated_at    TEXT NOT NULL,
    FOREIGN KEY (item_id) REFERENCES items(id) ON DELETE CASCADE
);

CREATE TABLE IF NOT EXISTS history (
    id        INTEGER PRIMARY KEY AUTOINCREMENT,
    item_id   INTEGER,
    target_kind TEXT NOT NULL,   -- item | response
    target_id INTEGER,
    action    TEXT NOT NULL,     -- create | update | delete | import
    field     TEXT,
    old_value TEXT,
    new_value TEXT,
    actor     TEXT NOT NULL DEFAULT '',
    ts        TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS attachments (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    item_id     INTEGER NOT NULL,
    response_id INTEGER,                    -- NULL 이면 요청에 직접 붙은 사진
    filename    TEXT NOT NULL,              -- 사용자가 올린 원래 이름
    stored      TEXT NOT NULL,              -- 실제 저장 파일명
    thumb       TEXT,                       -- 썸네일 파일명 (없으면 원본 사용)
    mime        TEXT NOT NULL DEFAULT 'image/png',
    size        INTEGER NOT NULL DEFAULT 0,
    uploader    TEXT NOT NULL DEFAULT '',
    created_at  TEXT NOT NULL
);

CREATE INDEX IF NOT EXISTS idx_att_item ON attachments(item_id);
CREATE INDEX IF NOT EXISTS idx_resp_item ON responses(item_id);
CREATE INDEX IF NOT EXISTS idx_hist_item ON history(item_id, id);
"""


def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
    return g.db


@app.teardown_appcontext
def close_db(exc):
    db = g.pop("db", None)
    if db is not None:
        db.close()


# 구버전 DB에 나중에 추가된 컬럼 (기존 reqlog.db 그대로 써도 자동 반영된다)
MIGRATIONS = [
    ("items", "confirmed_at", "TEXT"),
    ("items", "confirmed_by", "TEXT"),
]


def init_db():
    con = sqlite3.connect(DB_PATH)
    con.executescript(SCHEMA)
    for table, col, coltype in MIGRATIONS:
        cols = [r[1] for r in con.execute("PRAGMA table_info(%s)" % table)]
        if col not in cols:
            con.execute("ALTER TABLE %s ADD COLUMN %s %s" % (table, col, coltype))
            print(" [migrate] %s.%s 추가" % (table, col))
    con.commit()
    con.close()
    if not os.path.isdir(UPLOAD_DIR):
        os.makedirs(UPLOAD_DIR)


def now_str():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def actor_of(req):
    """작성자 이름은 헤더(X-Actor)로 받는다. 폐쇄망 내부용이라 별도 인증 없음."""
    name = req.headers.get("X-Actor", "") or ""
    try:
        name = name.encode("latin-1").decode("utf-8")
    except (UnicodeEncodeError, UnicodeDecodeError):
        pass
    return name.strip()[:50]


def missing_resp(row):
    """없는 건이면 404. 최종완료로 확정돼도 수정은 막지 않는다 —
    '고객이 확인했다'는 표시일 뿐, 잠금이 아니다."""
    if row is None:
        return jsonify({"ok": False, "error": "not found"}), 404
    return None


def item_row(db, iid):
    return db.execute("SELECT * FROM items WHERE id=?", (iid,)).fetchone()


def log_history(db, item_id, kind, target_id, action, field, old, new, actor):
    db.execute(
        "INSERT INTO history(item_id,target_kind,target_id,action,field,old_value,new_value,actor,ts)"
        " VALUES(?,?,?,?,?,?,?,?,?)",
        (item_id, kind, target_id, action, field,
         None if old is None else str(old), None if new is None else str(new),
         actor, now_str()),
    )


# ----------------------------------------------------------------------------
# 날짜 정규화 (엑셀 시리얼 / 2026.08.25 / 2026-08-25 / 20260825 전부 처리)
# ----------------------------------------------------------------------------
EXCEL_EPOCH = date(1899, 12, 30)


def norm_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s:
        return None
    # 엑셀 시리얼 날짜 (예: 46258 -> 2026-08-25)
    if re.fullmatch(r"\d{5}(\.\d+)?", s):
        try:
            n = int(float(s))
            if 20000 <= n <= 80000:
                return (EXCEL_EPOCH + timedelta(days=n)).strftime("%Y-%m-%d")
        except ValueError:
            pass
    m = re.search(r"(\d{4})[.\-/년]\s*(\d{1,2})[.\-/월]\s*(\d{1,2})", s)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return date(y, mo, d).strftime("%Y-%m-%d")
        except ValueError:
            return None
    if re.fullmatch(r"\d{8}", s):
        try:
            return datetime.strptime(s, "%Y%m%d").strftime("%Y-%m-%d")
        except ValueError:
            return None
    return None


def clean(v):
    if v is None:
        return ""
    return str(v).strip()


def norm_tags(v):
    parts = [p.strip() for p in re.split(r"[,;/|]", clean(v)) if p.strip()]
    seen, out = set(), []
    for p in parts:
        if p.upper() not in seen:
            seen.add(p.upper())
            out.append(p)
    return ", ".join(out)


def next_seq(db):
    row = db.execute("SELECT COALESCE(MAX(seq),0) AS m FROM items").fetchone()
    return int(row["m"]) + 1


# ----------------------------------------------------------------------------
# 조회
# ----------------------------------------------------------------------------
def fetch_items(db, args):
    where, params = [], []
    q = clean(args.get("q"))
    if q:
        where.append("(i.content LIKE ? OR i.requester LIKE ? OR i.target LIKE ? OR"
                     " EXISTS(SELECT 1 FROM responses r WHERE r.item_id=i.id AND"
                     " (r.content LIKE ? OR r.responder LIKE ?)))")
        like = "%" + q + "%"
        params += [like] * 5
    for col, key in (("category", "category"), ("status", "status")):
        vals = [v for v in args.getlist(key) if clean(v)]
        if vals:
            where.append("i.%s IN (%s)" % (col, ",".join("?" * len(vals))))
            params += vals
    tag = clean(args.get("target"))
    if tag:
        where.append("i.target LIKE ?")
        params.append("%" + tag + "%")
    conf = clean(args.get("confirmed"))
    if conf == "Y":
        where.append("i.confirmed_at IS NOT NULL")
    elif conf == "N":
        where.append("i.confirmed_at IS NULL")
    requester = clean(args.get("requester"))
    if requester:
        where.append("i.requester LIKE ?")
        params.append("%" + requester + "%")
    d_from, d_to = norm_date(args.get("from")), norm_date(args.get("to"))
    if d_from:
        where.append("IFNULL(i.request_date,'') >= ?")
        params.append(d_from)
    if d_to:
        where.append("IFNULL(i.request_date,'9999-12-31') <= ?")
        params.append(d_to)

    sql = "SELECT i.* FROM items i"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY i.seq DESC, i.id DESC"
    items = [dict(r) for r in db.execute(sql, params).fetchall()]
    if not items:
        return []
    ids = [it["id"] for it in items]
    rows = db.execute(
        "SELECT * FROM responses WHERE item_id IN (%s) ORDER BY id ASC"
        % ",".join("?" * len(ids)), ids).fetchall()
    bucket = {i: [] for i in ids}
    for r in rows:
        bucket[r["item_id"]].append(dict(r))

    # 첨부(사진) — 요청 직속 / 응답 소속 분리
    atts = db.execute(
        "SELECT id,item_id,response_id,filename,mime,size,uploader,created_at,"
        "(thumb IS NOT NULL) AS has_thumb FROM attachments WHERE item_id IN (%s)"
        " ORDER BY id ASC" % ",".join("?" * len(ids)), ids).fetchall()
    item_att = {i: [] for i in ids}
    resp_att = {}
    for a in atts:
        d = dict(a)
        d["is_image"] = ext_of(d["filename"]) in ALLOWED_IMG
        if d["response_id"]:
            resp_att.setdefault(d["response_id"], []).append(d)
        else:
            item_att[d["item_id"]].append(d)

    for it in items:
        it["responses"] = bucket.get(it["id"], [])
        for rp in it["responses"]:
            rp["attachments"] = resp_att.get(rp["id"], [])
        it["attachments"] = item_att.get(it["id"], [])
        it["tags"] = [t.strip() for t in it["target"].split(",") if t.strip()]
        it["att_total"] = len(it["attachments"]) + sum(
            len(r["attachments"]) for r in it["responses"])
    return items


# ----------------------------------------------------------------------------
# API
# ----------------------------------------------------------------------------
@app.route("/api/meta")
def api_meta():
    db = get_db()
    tags = set()
    for r in db.execute("SELECT target FROM items WHERE target<>''"):
        for t in r["target"].split(","):
            if t.strip():
                tags.add(t.strip())
    people = set()
    for r in db.execute("SELECT requester AS n FROM items WHERE requester<>''"):
        people.add(r["n"])
    for r in db.execute("SELECT responder AS n FROM responses WHERE responder<>''"):
        people.add(r["n"])
    counts = {row["status"]: row["c"] for row in
              db.execute("SELECT status, COUNT(*) c FROM items GROUP BY status")}
    total = db.execute("SELECT COUNT(*) c FROM items").fetchone()["c"]
    done = db.execute("SELECT COUNT(*) c FROM items"
                      " WHERE confirmed_at IS NOT NULL").fetchone()["c"]
    return jsonify({"categories": CATEGORIES, "statuses": STATUSES,
                    "tags": sorted(tags), "people": sorted(people),
                    "counts": counts, "total": total,
                    "confirmed": done, "open": total - done})


@app.route("/api/items")
def api_items():
    return jsonify({"items": fetch_items(get_db(), request.args)})


@app.route("/api/items", methods=["POST"])
def api_item_create():
    db, d = get_db(), request.get_json(force=True, silent=True) or {}
    actor = actor_of(request) or clean(d.get("requester"))
    seq = next_seq(db)
    cur = db.execute(
        "INSERT INTO items(seq,category,requester,request_date,content,target,status,"
        "applied_date,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?)",
        (seq,
         clean(d.get("category")) or "요청",
         clean(d.get("requester")),
         norm_date(d.get("request_date")) or date.today().strftime("%Y-%m-%d"),
         clean(d.get("content")),
         norm_tags(d.get("target")),
         clean(d.get("status")) or "대기",
         norm_date(d.get("applied_date")),
         now_str(), now_str()))
    iid = cur.lastrowid
    log_history(db, iid, "item", iid, "create", None, None,
                clean(d.get("content"))[:200], actor)
    db.commit()
    return jsonify({"ok": True, "id": iid, "seq": seq})


ITEM_FIELDS = {"category": "구분", "requester": "요청자", "request_date": "요청일",
               "content": "내용", "target": "대상", "status": "상태",
               "applied_date": "적용일"}


@app.route("/api/items/<int:iid>", methods=["PUT"])
def api_item_update(iid):
    db, d = get_db(), request.get_json(force=True, silent=True) or {}
    actor = actor_of(request)
    old = item_row(db, iid)
    missing = missing_resp(old)
    if missing:
        return missing
    sets, params = [], []
    for f in ITEM_FIELDS:
        if f not in d:
            continue
        if f in ("request_date", "applied_date"):
            nv = norm_date(d[f])
        elif f == "target":
            nv = norm_tags(d[f])
        else:
            nv = clean(d[f])
        ov = old[f]
        if (ov or "") != (nv or ""):
            sets.append("%s=?" % f)
            params.append(nv)
            log_history(db, iid, "item", iid, "update", ITEM_FIELDS[f], ov, nv, actor)
    changed = len(sets)
    if sets:
        sets.append("updated_at=?")
        params.append(now_str())
        params.append(iid)
        db.execute("UPDATE items SET %s WHERE id=?" % ",".join(sets), params)
        db.commit()
    return jsonify({"ok": True, "changed": changed})


@app.route("/api/items/<int:iid>", methods=["DELETE"])
def api_item_delete(iid):
    db = get_db()
    row = item_row(db, iid)
    missing = missing_resp(row)
    if missing:
        return missing
    log_history(db, iid, "item", iid, "delete", None, row["content"][:200], None,
                actor_of(request))
    purge_files(db, "SELECT stored,thumb FROM attachments WHERE item_id=?", (iid,))
    db.execute("DELETE FROM attachments WHERE item_id=?", (iid,))
    db.execute("DELETE FROM responses WHERE item_id=?", (iid,))
    db.execute("DELETE FROM items WHERE id=?", (iid,))
    db.commit()
    return jsonify({"ok": True})


def purge_files(db, sql, params):
    """삭제되는 레코드에 딸린 실제 이미지 파일을 지운다."""
    for r in db.execute(sql, params).fetchall():
        for f in (r["stored"], r["thumb"]):
            if not f:
                continue
            p = os.path.join(UPLOAD_DIR, os.path.basename(f))
            if os.path.isfile(p):
                try:
                    os.remove(p)
                except OSError:
                    pass


# ----------------------------------------------------------------------------
# 최종완료 확인 도장
#   고객(요청자)이 "확인했다"고 찍는 표시. 잠그지 않는다 —
#   찍힌 뒤에도 수정·응답·첨부 전부 그대로 되고, 취소도 된다.
#   누가 언제 찍고 취소했는지는 이력에 남는다.
# ----------------------------------------------------------------------------
@app.route("/api/items/<int:iid>/confirm", methods=["POST"])
def api_item_confirm(iid):
    db, d = get_db(), request.get_json(force=True, silent=True) or {}
    row = item_row(db, iid)
    if row is None:
        return jsonify({"ok": False, "error": "not found"}), 404
    if row["confirmed_at"]:
        return jsonify({"ok": False, "error": "이미 확인된 건이다"}), 400
    who = clean(d.get("by")) or actor_of(request)
    if not who:
        return jsonify({"ok": False, "error": "확인자 이름이 필요하다"}), 400
    ts = now_str()
    # 아직 진행중인 건이면 상태도 같이 적용완료로 올려준다 (반려·보류는 건드리지 않음)
    if row["status"] in ("대기", "검토중"):
        db.execute("UPDATE items SET status='적용완료',"
                   "applied_date=COALESCE(applied_date,?) WHERE id=?",
                   (date.today().strftime("%Y-%m-%d"), iid))
    db.execute("UPDATE items SET confirmed_at=?,confirmed_by=?,updated_at=? WHERE id=?",
               (ts, who, ts, iid))
    log_history(db, iid, "item", iid, "confirm", "최종완료", None,
                "%s 확인" % who, who)
    db.commit()
    return jsonify({"ok": True, "confirmed_at": ts, "confirmed_by": who})


@app.route("/api/items/<int:iid>/unconfirm", methods=["POST"])
def api_item_unconfirm(iid):
    db = get_db()
    row = item_row(db, iid)
    if row is None:
        return jsonify({"ok": False, "error": "not found"}), 404
    if not row["confirmed_at"]:
        return jsonify({"ok": False, "error": "확인 상태가 아니다"}), 400
    actor = actor_of(request)
    log_history(db, iid, "item", iid, "unconfirm", "최종완료 취소",
                "%s 확인(%s)" % (row["confirmed_by"], row["confirmed_at"]), None, actor)
    db.execute("UPDATE items SET confirmed_at=NULL,confirmed_by=NULL,updated_at=?"
               " WHERE id=?", (now_str(), iid))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/items/<int:iid>/responses", methods=["POST"])
def api_resp_create(iid):
    db, d = get_db(), request.get_json(force=True, silent=True) or {}
    missing = missing_resp(item_row(db, iid))
    if missing:
        return missing
    actor = actor_of(request) or clean(d.get("responder"))
    cur = db.execute(
        "INSERT INTO responses(item_id,responder,response_date,content,created_at,updated_at)"
        " VALUES(?,?,?,?,?,?)",
        (iid, clean(d.get("responder")) or actor,
         norm_date(d.get("response_date")) or date.today().strftime("%Y-%m-%d"),
         clean(d.get("content")), now_str(), now_str()))
    log_history(db, iid, "response", cur.lastrowid, "create", "응답", None,
                clean(d.get("content"))[:200], actor)
    db.execute("UPDATE items SET updated_at=? WHERE id=?", (now_str(), iid))
    db.commit()
    return jsonify({"ok": True, "id": cur.lastrowid})


@app.route("/api/responses/<int:rid>", methods=["PUT", "DELETE"])
def api_resp_edit(rid):
    db = get_db()
    old = db.execute("SELECT * FROM responses WHERE id=?", (rid,)).fetchone()
    if old is None:
        return jsonify({"ok": False, "error": "not found"}), 404
    missing = missing_resp(item_row(db, old["item_id"]))
    if missing:
        return missing
    actor = actor_of(request)
    if request.method == "DELETE":
        log_history(db, old["item_id"], "response", rid, "delete", "응답",
                    old["content"][:200], None, actor)
        purge_files(db, "SELECT stored,thumb FROM attachments WHERE response_id=?",
                    (rid,))
        db.execute("DELETE FROM attachments WHERE response_id=?", (rid,))
        db.execute("DELETE FROM responses WHERE id=?", (rid,))
        db.commit()
        return jsonify({"ok": True})
    d = request.get_json(force=True, silent=True) or {}
    labels = {"responder": "응답자", "response_date": "응답일", "content": "응답내용"}
    sets, params = [], []
    for f, label in labels.items():
        if f not in d:
            continue
        nv = norm_date(d[f]) if f == "response_date" else clean(d[f])
        if (old[f] or "") != (nv or ""):
            sets.append("%s=?" % f)
            params.append(nv)
            log_history(db, old["item_id"], "response", rid, "update", label,
                        old[f], nv, actor)
    changed = len(sets)
    if sets:
        sets.append("updated_at=?")
        params += [now_str(), rid]
        db.execute("UPDATE responses SET %s WHERE id=?" % ",".join(sets), params)
        db.commit()
    return jsonify({"ok": True, "changed": changed})


# ----------------------------------------------------------------------------
# 첨부(사진)
#   - 이미지 파일만 받는다. 엑셀에는 이미지가 아니라 '파일명/장수'만 나간다.
#   - 원본과 함께 브라우저에서 만든 썸네일을 같이 받아 목록을 가볍게 유지한다.
# ----------------------------------------------------------------------------
def ext_of(name):
    return (name.rsplit(".", 1)[-1].lower() if "." in (name or "") else "")


def save_blob(fs, ext):
    stored = "%s.%s" % (uuid.uuid4().hex, ext)
    path = os.path.join(UPLOAD_DIR, stored)
    fs.save(path)
    return stored, os.path.getsize(path)


@app.route("/api/items/<int:iid>/attachments", methods=["POST"])
def api_att_upload(iid):
    db = get_db()
    missing = missing_resp(item_row(db, iid))
    if missing:
        return missing
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "파일이 없다"}), 400
    fs = request.files["file"]
    ext = ext_of(fs.filename)
    if ext in BLOCKED_EXT:
        return jsonify({"ok": False,
                        "error": "실행/스크립트 파일(.%s)은 등록할 수 없다" % ext}), 400
    if ext not in ALLOWED_EXT:
        return jsonify({"ok": False,
                        "error": "허용되지 않는 형식(.%s)이다. 사진 또는 문서/압축 파일만 된다"
                                 % (ext or "확장자없음")}), 400

    rid = request.form.get("response_id") or None
    if rid:
        try:
            rid = int(rid)
        except ValueError:
            rid = None
        if rid and db.execute("SELECT 1 FROM responses WHERE id=? AND item_id=?",
                              (rid, iid)).fetchone() is None:
            return jsonify({"ok": False, "error": "응답을 찾을 수 없다"}), 404

    if not os.path.isdir(UPLOAD_DIR):
        os.makedirs(UPLOAD_DIR)
    stored, size = save_blob(fs, ext)

    thumb = None
    if "thumb" in request.files:
        try:
            thumb, _ = save_blob(request.files["thumb"], "jpg")
        except Exception:  # noqa: BLE001  썸네일 실패해도 원본은 살린다
            thumb = None

    actor = actor_of(request)
    cur = db.execute(
        "INSERT INTO attachments(item_id,response_id,filename,stored,thumb,mime,size,"
        "uploader,created_at) VALUES(?,?,?,?,?,?,?,?,?)",
        (iid, rid, os.path.basename(fs.filename)[:180], stored, thumb,
         MIME_BY_EXT.get(ext, "application/octet-stream"), size, actor, now_str()))
    log_history(db, iid, "attachment", cur.lastrowid, "create", "사진", None,
                os.path.basename(fs.filename)[:180], actor)
    db.execute("UPDATE items SET updated_at=? WHERE id=?", (now_str(), iid))
    db.commit()
    return jsonify({"ok": True, "id": cur.lastrowid, "size": size})


@app.route("/api/attachments/<int:aid>")
@app.route("/api/attachments/<int:aid>/<mode>")
def api_att_get(aid, mode=None):
    row = get_db().execute("SELECT * FROM attachments WHERE id=?", (aid,)).fetchone()
    if row is None:
        return jsonify({"ok": False, "error": "not found"}), 404
    use_thumb = (mode == "thumb") and row["thumb"]
    fname = row["thumb"] if use_thumb else row["stored"]
    path = os.path.join(UPLOAD_DIR, os.path.basename(fname))
    if not os.path.isfile(path):
        return jsonify({"ok": False, "error": "파일이 없다"}), 404

    is_img = ext_of(row["filename"]) in ALLOWED_IMG
    # 사진만 브라우저에서 바로 열고, 그 외 문서는 무조건 다운로드시킨다.
    resp = send_file(path,
                     mimetype="image/jpeg" if use_thumb else (
                         row["mime"] if is_img else "application/octet-stream"),
                     download_name=row["filename"],
                     as_attachment=(not is_img))
    resp.headers["X-Content-Type-Options"] = "nosniff"
    resp.headers["Cache-Control"] = "private, max-age=86400"
    return resp


@app.route("/api/attachments/<int:aid>", methods=["DELETE"])
def api_att_delete(aid):
    db = get_db()
    row = db.execute("SELECT * FROM attachments WHERE id=?", (aid,)).fetchone()
    if row is None:
        return jsonify({"ok": False, "error": "not found"}), 404
    missing = missing_resp(item_row(db, row["item_id"]))
    if missing:
        return missing
    for f in (row["stored"], row["thumb"]):
        if not f:
            continue
        p = os.path.join(UPLOAD_DIR, os.path.basename(f))
        if os.path.isfile(p):
            try:
                os.remove(p)
            except OSError:
                pass
    log_history(db, row["item_id"], "attachment", aid, "delete", "사진",
                row["filename"], None, actor_of(request))
    db.execute("DELETE FROM attachments WHERE id=?", (aid,))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/items/<int:iid>/history")
def api_history(iid):
    rows = get_db().execute(
        "SELECT * FROM history WHERE item_id=? ORDER BY id DESC LIMIT 300", (iid,)
    ).fetchall()
    return jsonify({"history": [dict(r) for r in rows]})


# ----------------------------------------------------------------------------
# 엑셀 내보내기
# ----------------------------------------------------------------------------
EXPORT_HEADERS = ["No.", "해당", "작성자/응답자", "작성날짜", "내용",
                  "대상(시스템/FAB)", "첨부파일", "적용여부", "적용날짜", "최종완료"]


def conf_label(it):
    """'2026-08-26 김윤환TL님' 형태. 비어 있으면 아직 고객 확인 전이다."""
    if not it["confirmed_at"]:
        return ""
    return "%s %s" % (it["confirmed_at"][:10], it["confirmed_by"] or "")


def att_label(atts):
    """엑셀에는 파일 자체가 아니라 '건수 + 파일명'만 찍는다.
    표만 받아본 사람도 뭐가 등록돼 있는지는 알 수 있게 하는 용도."""
    if not atts:
        return ""
    names = [a["filename"] for a in atts]
    return "%d건: %s" % (len(names), ", ".join(names))


@app.route("/api/export.xlsx")
def api_export():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    items = fetch_items(get_db(), request.args)
    items = sorted(items, key=lambda x: (x["seq"] or 0, x["id"]))

    wb = Workbook()
    ws = wb.active
    ws.title = "요청이력"

    head_fill = PatternFill("solid", fgColor="1F3864")
    head_font = Font(color="FFFFFF", bold=True, size=10)
    req_fill = PatternFill("solid", fgColor="EDF2FA")
    done_fill = PatternFill("solid", fgColor="E7F6EC")   # 고객 확인 끝난 건
    done_font = Font(bold=True, color="166B38")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.append(EXPORT_HEADERS)
    for c in ws[1]:
        c.fill, c.font, c.alignment, c.border = head_fill, head_font, center, border

    for it in items:
        ws.append([it["seq"], it["category"], it["requester"], it["request_date"],
                   it["content"], it["target"], att_label(it["attachments"]),
                   it["status"], it["applied_date"] or "", conf_label(it)])
        r = ws.max_row
        done = bool(it["confirmed_at"])
        for c in ws[r]:
            c.fill, c.border, c.alignment = (done_fill if done else req_fill), border, wrap
        ws.cell(row=r, column=1).alignment = center
        ws.cell(row=r, column=2).alignment = center
        ws.cell(row=r, column=8).alignment = center
        ws.cell(row=r, column=10).alignment = center
        if done:
            ws.cell(row=r, column=10).font = done_font
        for rp in it["responses"]:
            ws.append(["", "확인", rp["responder"], rp["response_date"],
                       rp["content"], "", att_label(rp["attachments"]), "", "", ""])
            rr = ws.max_row
            for c in ws[rr]:
                c.border, c.alignment = border, wrap
            ws.cell(row=rr, column=2).alignment = center

    widths = [6, 8, 16, 13, 58, 18, 28, 11, 13, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:J%d" % max(ws.max_row, 1)

    # 요약 시트
    ws2 = wb.create_sheet("요약")
    ws2.append(["No.", "구분", "요청자", "요청일", "내용", "대상", "상태",
                "적용일", "최종완료", "응답수", "첨부수", "최근응답"])
    for c in ws2[1]:
        c.fill, c.font, c.alignment, c.border = head_fill, head_font, center, border
    for it in items:
        last = it["responses"][-1]["content"] if it["responses"] else ""
        ws2.append([it["seq"], it["category"], it["requester"], it["request_date"],
                    it["content"], it["target"], it["status"],
                    it["applied_date"] or "", conf_label(it), len(it["responses"]),
                    it["att_total"], last])
        if it["confirmed_at"]:
            ws2.cell(row=ws2.max_row, column=9).font = done_font
    for i, w in enumerate([6, 8, 14, 13, 46, 18, 10, 13, 22, 8, 8, 46], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    for row in ws2.iter_rows(min_row=2):
        for c in row:
            c.border, c.alignment = border, wrap
    ws2.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = "요청이력_%s.xlsx" % datetime.now().strftime("%Y%m%d_%H%M")
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument."
                              "spreadsheetml.sheet")


# ----------------------------------------------------------------------------
# 엑셀/CSV 가져오기
#   기존 표 형식: No. | 해당 | 작성자/응답자 | 작성날짜 | 내용 | 적용여부 | 적용날짜
#   No.가 있으면 새 요청, 비어있고 '확인/응답'이면 직전 요청의 응답으로 붙인다.
# ----------------------------------------------------------------------------
RESP_KINDS = {"확인", "응답", "답변", "회신", "코멘트", "댓글"}

# 헤더 이름 → 내부 필드. 기존 표와 이 시스템이 뽑은 엑셀 둘 다 받는다.
HEADER_MAP = [
    ("no", ("no", "no.", "번호")),
    ("kind", ("해당", "구분")),
    ("who", ("작성자", "응답자", "요청자", "작성자/응답자")),
    ("date", ("작성날짜", "요청일", "작성일", "날짜")),
    ("content", ("내용",)),
    ("target", ("대상",)),
    ("attach", ("첨부", "사진", "첨부파일")),
    ("applied", ("적용여부", "상태")),
    ("applied_date", ("적용날짜", "적용일")),
    ("confirmed", ("최종완료", "고객확인", "확인완료")),
]
LEGACY_COLS = ["no", "kind", "who", "date", "content", "applied", "applied_date"]


def build_colmap(rows):
    """헤더 행을 찾아 컬럼 위치를 잡는다. 못 찾으면 기존 표 순서로 간주."""
    for row in rows[:6]:
        cells = [clean(c) for c in row]
        joined = " ".join(cells)
        if "내용" not in joined:
            continue
        cmap = {}
        for idx, cell_txt in enumerate(cells):
            low = cell_txt.replace(" ", "").lower()
            if not low:
                continue
            for field, keys in HEADER_MAP:
                if field in cmap:
                    continue
                if any(k in low for k in keys):
                    cmap[field] = idx
                    break
        if "content" in cmap:
            return cmap, rows.index(row)
    return {f: i for i, f in enumerate(LEGACY_COLS)}, -1


def read_rows(fs):
    name = (fs.filename or "").lower()
    raw = fs.read()
    if name.endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb.active
        return [list(r) for r in ws.iter_rows(values_only=True)]
    text = None
    for enc in ("utf-8-sig", "cp949", "euc-kr", "utf-8"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = raw.decode("utf-8", errors="replace")
    delim = "\t" if text.count("\t") > text.count(",") else ","
    return [row for row in csv.reader(io.StringIO(text), delimiter=delim)]


def cell(row, i):
    return row[i] if i < len(row) and row[i] is not None else ""


@app.route("/api/import", methods=["POST"])
def api_import():
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "파일이 없다"}), 400
    db, actor = get_db(), actor_of(request)
    try:
        rows = read_rows(request.files["file"])
    except Exception as e:  # noqa: BLE001
        return jsonify({"ok": False, "error": "파일 파싱 실패: %s" % e}), 400

    cmap, head_at = build_colmap(rows)
    n_item = n_resp = 0
    last_id = None
    seq = next_seq(db)
    for ridx, row in enumerate(rows):
        if ridx <= head_at:
            continue
        if not any(clean(c) for c in row):
            continue

        def col(field):
            i = cmap.get(field)
            return clean(cell(row, i)) if i is not None else ""

        no, kind, who = col("no"), col("kind"), col("who")
        d1, content = col("date"), col("content")
        target, applied, applied_d = col("target"), col("applied"), col("applied_date")
        conf_raw = col("confirmed")
        if no.lower().startswith("no") or content == "내용":
            continue  # 헤더 행
        is_resp = (not no) and (kind in RESP_KINDS or bool(who))
        the_date = norm_date(d1) or norm_date(applied) or norm_date(applied_d)
        if is_resp and last_id:
            db.execute(
                "INSERT INTO responses(item_id,responder,response_date,content,"
                "created_at,updated_at) VALUES(?,?,?,?,?,?)",
                (last_id, who, the_date, content, now_str(), now_str()))
            n_resp += 1
            continue
        if not (content or who):
            continue
        cat = kind if kind in CATEGORIES else "요청"
        st = "적용완료" if (applied and applied not in STATUSES and
                        norm_date(applied)) else (
            applied if applied in STATUSES else "대기")
        # '최종완료' 컬럼: "2026-08-26 김윤환TL님" 형태에서 날짜와 확인자를 뽑는다
        c_at = c_by = None
        if conf_raw:
            c_date = norm_date(conf_raw[:12]) or norm_date(conf_raw)
            if c_date:
                c_at = c_date + " 00:00:00"
                c_by = re.sub(r"[\d.\-/:]+", " ", conf_raw).strip() or None
            elif conf_raw.upper() in ("O", "Y", "YES", "완료", "확인"):
                c_at = (the_date or date.today().strftime("%Y-%m-%d")) + " 00:00:00"
        cur = db.execute(
            "INSERT INTO items(seq,category,requester,request_date,content,target,"
            "status,applied_date,confirmed_at,confirmed_by,created_at,updated_at)"
            " VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
            (seq, cat, who, the_date, content, norm_tags(target), st,
             norm_date(applied_d) or norm_date(applied), c_at, c_by,
             now_str(), now_str()))
        last_id = cur.lastrowid
        log_history(db, last_id, "item", last_id, "import", None, None,
                    content[:200], actor)
        seq += 1
        n_item += 1
    db.commit()
    note = ""
    if "attach" in cmap:
        note = "엑셀에는 첨부 파일명만 들어있다. 파일 자체는 복원되지 않으니 웹에서 다시 등록해라."
    return jsonify({"ok": True, "items": n_item, "responses": n_resp, "note": note})


# ----------------------------------------------------------------------------
# 화면
# ----------------------------------------------------------------------------
@app.route("/")
def index():
    return Response(INDEX_HTML, mimetype="text/html; charset=utf-8")


INDEX_HTML = r"""<!doctype html>
<html lang="ko">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>요청/응답 관리</title>
<style>
*{box-sizing:border-box}
:root{
  --bg:#f4f6fa; --panel:#fff; --line:#dde3ec; --text:#1a2233; --muted:#6b7688;
  --accent:#1f3864; --accent2:#2f6fd0; --warn:#c0392b; --ok:#1e8449;
}
body{margin:0;background:var(--bg);color:var(--text);
  font-family:"Malgun Gothic","맑은 고딕",AppleSDGothicNeo-Regular,"Apple SD Gothic Neo",
  "Noto Sans KR",Dotum,sans-serif;font-size:13.5px;line-height:1.55}
header{background:var(--accent);color:#fff;padding:10px 16px;display:flex;
  align-items:center;gap:12px;flex-wrap:wrap;position:sticky;top:0;z-index:20}
header h1{font-size:16px;margin:0;font-weight:700;letter-spacing:-.3px}
header .sp{flex:1}
header input{border:0;border-radius:4px;padding:5px 8px;font-size:13px;width:130px}
button{font-family:inherit;font-size:13px;border:1px solid var(--line);
  background:#fff;color:var(--text);border-radius:4px;padding:5px 11px;cursor:pointer}
button:hover{background:#eef2f8}
button.p{background:var(--accent2);border-color:var(--accent2);color:#fff}
button.p:hover{filter:brightness(1.08)}
button.d{color:var(--warn);border-color:#eccfcb}
button.sm{padding:3px 8px;font-size:12px}
.wrap{max-width:1240px;margin:0 auto;padding:14px 16px 60px}
.bar{background:var(--panel);border:1px solid var(--line);border-radius:6px;
  padding:10px 12px;display:flex;gap:8px;flex-wrap:wrap;align-items:center;margin-bottom:12px}
.bar input,.bar select{font-family:inherit;font-size:13px;padding:5px 7px;
  border:1px solid var(--line);border-radius:4px;background:#fff}
.bar input[type=text]{min-width:190px}
.stats{color:var(--muted);font-size:12.5px;margin:0 0 8px 2px}
.card{background:var(--panel);border:1px solid var(--line);border-radius:6px;
  margin-bottom:10px;overflow:hidden}
.card.on{border-color:var(--accent2);box-shadow:0 1px 6px rgba(47,111,208,.13)}
.card.done{border-left:4px solid var(--ok)}
.stamp{display:inline-flex;align-items:center;gap:5px;border:1.5px solid var(--ok);
  color:var(--ok);background:#eaf7ee;border-radius:4px;padding:1px 8px;font-size:11.5px;
  font-weight:700;white-space:nowrap}
.confbar{display:flex;gap:9px;align-items:center;flex-wrap:wrap;margin-top:9px;
  border-top:1px solid var(--line);padding-top:9px}
.confbar .msg{color:var(--ok);font-size:12.5px;font-weight:600}
button.ok{background:var(--ok);border-color:var(--ok);color:#fff;font-weight:600}
button.ok:hover{filter:brightness(1.08)}
.chead{display:flex;gap:10px;padding:11px 13px;cursor:pointer;align-items:flex-start}
.chead:hover{background:#fafbfe}
.no{font-weight:700;color:var(--muted);min-width:34px;font-size:13px;padding-top:1px}
.main{flex:1;min-width:0}
.titleline{display:flex;gap:6px;flex-wrap:wrap;align-items:center;margin-bottom:3px}
.ctext{white-space:pre-wrap;word-break:break-all}
.meta{color:var(--muted);font-size:12px;margin-top:4px}
.badge{display:inline-block;padding:1px 7px;border-radius:10px;font-size:11.5px;
  border:1px solid var(--line);background:#f2f5fa;color:#42506b;white-space:nowrap}
.b-요청{background:#e8f0fd;border-color:#c5d9f7;color:#1c4f9c}
.b-제안{background:#eaf7ee;border-color:#c6e6d1;color:#1e7a45}
.b-확인{background:#f3eefc;border-color:#ddd0f2;color:#5b3a9e}
.b-이슈{background:#fdecea;border-color:#f5cbc6;color:#b03a2e}
.s-대기{background:#f2f4f7;color:#5a6577}
.s-검토중{background:#fff6e0;border-color:#f2dfae;color:#96690b}
.s-적용완료{background:#e7f6ec;border-color:#bfe4cb;color:#166b38}
.s-보류{background:#eef0f4;border-color:#d7dbe3;color:#5a6577}
.s-반려{background:#fdecea;border-color:#f5cbc6;color:#b03a2e}
.tag{background:#eef2f7;border:1px solid #dbe3ee;color:#3d4a5f;border-radius:3px;
  padding:1px 6px;font-size:11.5px}
.body{border-top:1px solid var(--line);padding:11px 13px;background:#fbfcfe}
.resp{border-left:3px solid #cbd7ea;padding:6px 0 6px 10px;margin:8px 0}
.resp .who{font-weight:600;font-size:12.5px}
.resp .rtext{white-space:pre-wrap;word-break:break-all;margin-top:2px}
.row{display:flex;gap:8px;flex-wrap:wrap;align-items:center;margin-top:8px}
.row input,.row select,.row textarea{font-family:inherit;font-size:13px;padding:5px 7px;
  border:1px solid var(--line);border-radius:4px;background:#fff}
.row textarea{flex:1;min-width:260px;min-height:34px;resize:vertical}
.atts{display:flex;gap:7px;flex-wrap:wrap;margin-top:7px}
.thumb{position:relative;width:104px;height:78px;border:1px solid var(--line);
  border-radius:4px;overflow:hidden;background:#eef1f6;cursor:zoom-in}
.thumb img{width:100%;height:100%;object-fit:cover;display:block}
.thumb .x{position:absolute;top:2px;right:2px;background:rgba(20,26,38,.72);color:#fff;
  border:0;border-radius:3px;padding:0 5px;font-size:12px;line-height:17px;cursor:pointer;
  display:none}
.thumb:hover .x{display:block}
.thumb .fn{position:absolute;left:0;right:0;bottom:0;background:rgba(20,26,38,.62);
  color:#fff;font-size:10.5px;padding:1px 4px;white-space:nowrap;overflow:hidden;
  text-overflow:ellipsis}
.files{display:flex;flex-direction:column;gap:4px;margin-top:7px}
.file{display:flex;align-items:center;gap:8px;border:1px solid var(--line);
  border-radius:4px;padding:4px 8px;background:#fff;max-width:560px}
.file .ic{font-size:10.5px;font-weight:700;color:#fff;background:#5a6577;border-radius:3px;
  padding:1px 5px;min-width:42px;text-align:center;letter-spacing:.2px}
.file .ic.xls{background:#1e7a45}.file .ic.doc{background:#2f6fd0}
.file .ic.ppt{background:#c0562b}.file .ic.pdf{background:#b03a2e}
.file .ic.hwp{background:#1f6f9e}.file .ic.zip{background:#7a6a3a}
.file a{color:var(--accent2);text-decoration:none;flex:1;min-width:0;overflow:hidden;
  text-overflow:ellipsis;white-space:nowrap}
.file a:hover{text-decoration:underline}
.file .sz{color:var(--muted);font-size:11.5px;white-space:nowrap}
.attbar{display:flex;gap:8px;align-items:center;flex-wrap:wrap;margin-top:7px;
  font-size:12px;color:var(--muted)}
.drop{outline:2px dashed var(--accent2);outline-offset:-4px}
.lightbox{position:fixed;inset:0;background:rgba(12,16,24,.88);z-index:80;display:flex;
  flex-direction:column;align-items:center;justify-content:center;padding:20px}
.lightbox img{max-width:96vw;max-height:86vh;object-fit:contain;
  background:#fff;border-radius:4px}
.lightbox .cap{color:#e6ebf5;font-size:12.5px;margin-top:9px}
.hist{margin-top:10px;border-top:1px dashed var(--line);padding-top:8px;
  font-size:12px;color:var(--muted)}
.hist div{padding:1px 0}
.modal{position:fixed;inset:0;background:rgba(20,26,38,.45);display:flex;
  align-items:flex-start;justify-content:center;padding:40px 14px;z-index:50}
.modal .box{background:#fff;border-radius:8px;width:min(680px,100%);padding:18px 20px;
  max-height:86vh;overflow:auto}
.modal h2{margin:0 0 12px;font-size:16px}
.f{margin-bottom:10px}
.f label{display:block;font-size:12.5px;color:var(--muted);margin-bottom:3px}
.f input,.f select,.f textarea{width:100%;font-family:inherit;font-size:13.5px;
  padding:7px 8px;border:1px solid var(--line);border-radius:4px;background:#fff}
.f textarea{min-height:100px;resize:vertical}
.g2{display:flex;gap:10px}.g2>*{flex:1}
.empty{text-align:center;color:var(--muted);padding:44px 10px}
.toast{position:fixed;left:50%;transform:translateX(-50%);bottom:24px;background:#1a2233;
  color:#fff;padding:9px 16px;border-radius:5px;z-index:99;font-size:13px}
</style>
</head>
<body>
<header>
  <h1>요청 / 응답 관리 시스템</h1>
  <span class="sp"></span>
  <label style="font-size:12.5px">작성자
    <input id="actor" placeholder="이름 입력" title="기록에 남을 내 이름">
  </label>
  <button class="p" onclick="openItem()">+ 새 요청</button>
  <button onclick="doExport()">엑셀 다운로드</button>
  <button onclick="document.getElementById('fileup').click()">엑셀 업로드</button>
  <input type="file" id="fileup" accept=".xlsx,.xlsm,.csv,.tsv" style="display:none"
         onchange="doImport(this)">
</header>

<div class="wrap">
  <div class="bar">
    <input type="text" id="q" placeholder="내용·작성자·태그 검색" oninput="deb()">
    <select id="fcat" onchange="load()"><option value="">구분 전체</option></select>
    <select id="fst" onchange="load()"><option value="">상태 전체</option></select>
    <select id="fconf" onchange="load()">
      <option value="">최종완료 전체</option>
      <option value="N">미확인 (진행중)</option>
      <option value="Y">최종완료</option>
    </select>
    <select id="ftag" onchange="load()"><option value="">대상 전체</option></select>
    <input type="date" id="fd1" onchange="load()" title="요청일 시작">
    <span style="color:var(--muted)">~</span>
    <input type="date" id="fd2" onchange="load()" title="요청일 종료">
    <button class="sm" onclick="resetF()">초기화</button>
  </div>
  <div class="stats" id="stats"></div>
  <div id="list"></div>
</div>

<div id="modal"></div>
<script>
var CATS=[],STS=[],TAGS=[],ITEMS=[],OPEN={};
function $(id){return document.getElementById(id)}
function esc(s){return (s==null?'':String(s)).replace(/[&<>"']/g,function(c){
  return {'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c]})}
function actor(){return $('actor').value.trim()}
function toast(m){var d=document.createElement('div');d.className='toast';d.textContent=m;
  document.body.appendChild(d);setTimeout(function(){d.remove()},2200)}
function api(url,opt){opt=opt||{};opt.headers=opt.headers||{};
  opt.headers['X-Actor']=encodeHeader(actor());
  if(opt.body&&typeof opt.body!=='string'){opt.headers['Content-Type']='application/json';
    opt.body=JSON.stringify(opt.body)}
  return fetch(url,opt).then(function(r){return r.json()})}
function encodeHeader(s){try{return unescape(encodeURIComponent(s))}catch(e){return ''}}
function qs(){
  var p=new URLSearchParams();
  if($('q').value.trim())p.set('q',$('q').value.trim());
  if($('fcat').value)p.set('category',$('fcat').value);
  if($('fst').value)p.set('status',$('fst').value);
  if($('fconf').value)p.set('confirmed',$('fconf').value);
  if($('ftag').value)p.set('target',$('ftag').value);
  if($('fd1').value)p.set('from',$('fd1').value);
  if($('fd2').value)p.set('to',$('fd2').value);
  return p.toString()}
var t=null;function deb(){clearTimeout(t);t=setTimeout(load,250)}
function resetF(){$('q').value='';$('fcat').value='';$('fst').value='';$('ftag').value='';
  $('fconf').value='';$('fd1').value='';$('fd2').value='';load()}

function meta(){return api('/api/meta').then(function(m){
  CATS=m.categories;STS=m.statuses;TAGS=m.tags;
  fill('fcat','구분 전체',CATS);fill('fst','상태 전체',STS);fill('ftag','대상 전체',TAGS);
  var s=[];for(var i=0;i<STS.length;i++){var c=m.counts[STS[i]]||0;if(c)s.push(STS[i]+' '+c)}
  $('stats').innerHTML='전체 '+m.total+'건  ·  <b style="color:var(--ok)">최종완료 '+
    m.confirmed+'</b>  ·  미확인 '+m.open+
    (s.length?'  ·  '+esc(s.join('  ·  ')):'')})}
function fill(id,all,arr){var el=$(id),cur=el.value;
  el.innerHTML='<option value="">'+all+'</option>'+arr.map(function(v){
    return '<option>'+esc(v)+'</option>'}).join('');el.value=cur}

function load(){return api('/api/items?'+qs()).then(function(d){ITEMS=d.items;render()})}

function render(){
  var h=$('list');
  if(!ITEMS.length){h.innerHTML='<div class="empty">기록이 없다. [+ 새 요청] 또는 [엑셀 업로드]로 시작해라.</div>';return}
  h.innerHTML=ITEMS.map(card).join('')}

function card(it){
  var op=!!OPEN[it.id];
  var done=!!it.confirmed_at;
  var tags=it.tags.map(function(t){return '<span class="tag">'+esc(t)+'</span>'}).join(' ');
  var s='<div class="card'+(op?' on':'')+(done?' done':'')+'" id="c'+it.id+'">'+
   '<div class="chead" onclick="tog('+it.id+')">'+
     '<div class="no">'+esc(it.seq)+'</div><div class="main">'+
       '<div class="titleline">'+
         '<span class="badge b-'+esc(it.category)+'">'+esc(it.category)+'</span>'+
         '<span class="badge s-'+esc(it.status)+'">'+esc(it.status)+'</span>'+
         (done?'<span class="stamp">최종완료 '+esc((it.confirmed_at||'').slice(0,10))+
               ' · '+esc(it.confirmed_by||'')+'</span>':'')+tags+
       '</div>'+
       '<div class="ctext">'+esc(it.content)+'</div>'+
       '<div class="meta">'+esc(it.requester||'-')+' · '+esc(it.request_date||'-')+
         ' · 응답 '+it.responses.length+'건'+
         (it.att_total?' · 첨부 '+it.att_total+'건':'')+
         (it.applied_date?' · 적용 '+esc(it.applied_date):'')+'</div>'+
     '</div></div>';
  if(op){
    s+='<div class="body" id="body'+it.id+'" ondragover="dragOn(event,this)"'+
       ' ondragleave="dragOff(this)" ondrop="dropFiles(event,this,'+it.id+',0)">';
    s+=attBlock(it.attachments);
    s+='<div class="attbar">'+
       '<button class="sm" onclick="pick('+it.id+',0)">파일 추가</button>'+
       '<span>사진·엑셀·PDF·한글 등. 이 영역에 <b>끌어다 놓거나</b> 캡처를 '+
       '<b>Ctrl+V</b> 해도 등록된다. (엑셀 다운로드에는 파일명만 나간다)</span></div>';
    s+=it.responses.map(function(r){
      return '<div class="resp"><div class="who">'+esc(r.responder||'-')+
        ' <span style="color:var(--muted);font-weight:400">'+esc(r.response_date||'')+
        '</span></div><div class="rtext">'+esc(r.content)+'</div>'+
        attBlock(r.attachments)+
        '<div style="margin-top:3px"><button class="sm" onclick="editResp('+r.id+
        ')">수정</button> <button class="sm" onclick="pick('+it.id+','+r.id+
        ')">파일 추가</button> <button class="sm d" onclick="delResp('+r.id+
        ')">삭제</button></div></div>'}).join('');
    s+='<div class="row"><textarea id="rc'+it.id+'" placeholder="응답 / 확인 내용 입력"></textarea>'+
       '<input type="date" id="rd'+it.id+'" value="'+today()+'">'+
       '<button class="p" onclick="addResp('+it.id+')">응답 등록</button></div>';
    s+='<div class="row" style="border-top:1px solid var(--line);padding-top:9px">'+
       '상태 <select onchange="quick('+it.id+',\'status\',this.value)">'+
       STS.map(function(v){return '<option'+(v===it.status?' selected':'')+'>'+esc(v)+'</option>'}).join('')+
       '</select>'+
       '적용일 <input type="date" value="'+esc(it.applied_date||'')+
       '" onchange="quick('+it.id+',\'applied_date\',this.value)">'+
       '<span class="sp" style="flex:1"></span>'+
       '<button class="sm" onclick="openItem('+it.id+')">수정</button>'+
       '<button class="sm" onclick="showHist('+it.id+')">이력</button>'+
       '<button class="sm d" onclick="delItem('+it.id+')">삭제</button></div>';
    s+='<div class="confbar">'+(done
      ? '<span class="msg">고객 확인 완료 — '+esc(it.confirmed_by||'')+' · '+
        esc(it.confirmed_at||'')+'</span>'+
        '<span style="flex:1"></span>'+
        '<button class="sm" onclick="unconfirm('+it.id+')">최종완료 취소</button>'
      : '<span style="color:var(--muted);font-size:12.5px">처리 끝났으면 요청자가 여기서 '+
        '확인 도장을 찍는다. (찍어도 계속 수정·응답 가능하다)</span>'+
        '<span style="flex:1"></span>'+
        '<button class="ok" onclick="confirmItem('+it.id+')">최종완료 확인</button>')+
      '</div>';
    s+='<div class="hist" id="h'+it.id+'" style="display:none"></div>';
    s+='</div>'}
  return s+'</div>'}

function confirmItem(id){
  var who=actor();
  if(!who){toast('상단에 이름부터 입력해라 — 누가 확인했는지 남아야 한다');return}
  if(!confirm('["'+who+'" 이름으로 최종완료 확인을 찍는다]\n\n'+
              '잠기는 건 아니다. 이후에도 수정·응답·첨부 다 되고, 취소도 된다.\n\n'+
              '진행할까?'))return;
  api('/api/items/'+id+'/confirm',{method:'POST',body:{by:who}}).then(function(d){
    if(!d.ok){toast(d.error||'실패');return}
    toast('최종완료 확인됨');return Promise.all([meta(),load()])})}

function unconfirm(id){
  if(!actor()){toast('상단에 이름부터 입력해라');return}
  if(!confirm('최종완료 확인을 취소한다. 이력에 남는다. 진행할까?'))return;
  api('/api/items/'+id+'/unconfirm',{method:'POST',body:{}}).then(function(d){
    if(!d.ok){toast(d.error||'실패');return}
    toast('확인 취소됨');return Promise.all([meta(),load()])})}

/* ---------- 첨부 (사진 = 썸네일 / 문서 = 파일칩) ---------- */
function fsize(n){
  if(n<1024)return n+'B';
  if(n<1024*1024)return (n/1024).toFixed(0)+'KB';
  return (n/1024/1024).toFixed(1)+'MB'}
function fext(name){var i=(name||'').lastIndexOf('.');
  return i<0?'FILE':name.slice(i+1).toUpperCase()}
function icls(e){e=e.toLowerCase();
  if(e.indexOf('xls')===0||e==='csv'||e==='tsv')return 'xls';
  if(e.indexOf('doc')===0)return 'doc';
  if(e.indexOf('ppt')===0)return 'ppt';
  if(e==='pdf')return 'pdf';
  if(e.indexOf('hwp')===0)return 'hwp';
  if(['zip','7z','tar','gz','tgz'].indexOf(e)>=0)return 'zip';
  return ''}

function attBlock(list){
  if(!list||!list.length)return '';
  var imgs=list.filter(function(a){return a.is_image});
  var docs=list.filter(function(a){return !a.is_image});
  var s='';
  if(imgs.length){
    s+='<div class="atts">'+imgs.map(function(a){
      var src='/api/attachments/'+a.id+(a.has_thumb?'/thumb':'');
      return '<div class="thumb" onclick="viewAtt('+a.id+',this)">'+
        '<img src="'+src+'" alt="'+esc(a.filename)+'" loading="lazy">'+
        '<button class="x" title="삭제" onclick="event.stopPropagation();delAtt('+a.id+')">x</button>'+
        '<div class="fn" title="'+esc(a.filename)+'">'+esc(a.filename)+'</div></div>'
    }).join('')+'</div>'}
  if(docs.length){
    s+='<div class="files">'+docs.map(function(a){
      var e=fext(a.filename);
      return '<div class="file"><span class="ic '+icls(e)+'">'+esc(e.slice(0,4))+'</span>'+
        '<a href="/api/attachments/'+a.id+'" title="'+esc(a.filename)+
        '" download>'+esc(a.filename)+'</a>'+
        '<span class="sz">'+fsize(a.size)+'</span>'+
        '<button class="sm d" onclick="delAtt('+a.id+')">삭제</button></div>'
    }).join('')+'</div>'}
  return s}

function viewAtt(id,el){
  var name=el.querySelector('.fn').textContent;
  var d=document.createElement('div');d.className='lightbox';
  d.innerHTML='<img src="/api/attachments/'+id+'" alt=""><div class="cap">'+esc(name)+
    ' — 클릭하면 닫는다</div>';
  d.onclick=function(){d.remove()};document.body.appendChild(d)}

function delAtt(id){if(!confirm('이 첨부 삭제할까?'))return;
  api('/api/attachments/'+id,{method:'DELETE'}).then(function(){toast('첨부 삭제');
    return load()})}

/* 썸네일은 사진일 때만 만든다 (문서는 그대로 올린다) */
function makeThumb(file){return new Promise(function(res){
  if(!file.type||file.type.indexOf('image/')!==0){res(null);return}
  var url=URL.createObjectURL(file),img=new Image();
  img.onload=function(){
    var mx=360,s=Math.min(1,mx/Math.max(img.width,img.height));
    var c=document.createElement('canvas');
    c.width=Math.max(1,Math.round(img.width*s));c.height=Math.max(1,Math.round(img.height*s));
    c.getContext('2d').drawImage(img,0,0,c.width,c.height);
    URL.revokeObjectURL(url);
    if(c.toBlob){c.toBlob(function(b){res(b)},'image/jpeg',0.8)}else{res(null)}};
  img.onerror=function(){URL.revokeObjectURL(url);res(null)};
  img.src=url})}

function upOne(itemId,respId,file,name){
  return makeThumb(file).then(function(tb){
    var fd=new FormData();
    fd.append('file',file,name||file.name||'capture.png');
    if(tb)fd.append('thumb',tb,'t.jpg');
    if(respId)fd.append('response_id',respId);
    return fetch('/api/items/'+itemId+'/attachments',
      {method:'POST',body:fd,headers:{'X-Actor':encodeHeader(actor())}})
      .then(function(r){return r.json()})})}

function upload(itemId,respId,files){
  if(!files||!files.length)return;
  if(!actor()){toast('상단에 작성자 이름부터 입력해라');return}
  var arr=[],i;
  for(i=0;i<files.length;i++)arr.push(files[i]);
  toast('첨부 '+arr.length+'개 올리는 중...');
  var chain=Promise.resolve(),bad=[];
  arr.forEach(function(f){chain=chain.then(function(){
    return upOne(itemId,respId,f).then(function(d){
      if(!d||!d.ok)bad.push((f.name||'?')+' — '+((d&&d.error)||'실패'))})})});
  chain.then(function(){
    if(bad.length)alert('아래 파일은 등록 안 됐다:\n\n'+bad.join('\n'));
    else toast('첨부 등록 완료');
    return load()})}

function pick(itemId,respId){
  var el=document.createElement('input');el.type='file';el.multiple=true;
  el.onchange=function(){upload(itemId,respId,el.files)};el.click()}

function dragOn(e,el){e.preventDefault();el.classList.add('drop')}
function dragOff(el){el.classList.remove('drop')}
function dropFiles(e,el,itemId,respId){e.preventDefault();el.classList.remove('drop');
  upload(itemId,respId,e.dataTransfer.files)}

/* 캡처 붙여넣기: 펼쳐진 카드가 하나일 때 그 카드에 붙는다 */
document.addEventListener('paste',function(e){
  var ids=Object.keys(OPEN).filter(function(k){return OPEN[k]});
  if(ids.length!==1)return;
  var items=(e.clipboardData||{}).items||[],fs=[],i;
  for(i=0;i<items.length;i++){
    if(items[i].type&&items[i].type.indexOf('image')===0){
      var f=items[i].getAsFile();if(f)fs.push(f)}}
  if(!fs.length)return;
  e.preventDefault();
  var stamp=new Date().toISOString().slice(0,19).replace(/[:T]/g,'');
  if(!actor()){toast('상단에 작성자 이름부터 입력해라');return}
  toast('붙여넣은 캡처 등록 중...');
  var chain=Promise.resolve();
  fs.forEach(function(f,n){chain=chain.then(function(){
    return upOne(ids[0],0,f,'capture_'+stamp+(n?'_'+n:'')+'.png')})});
  chain.then(function(){toast('캡처 등록 완료');return load()})});

function today(){var d=new Date();return d.getFullYear()+'-'+
  ('0'+(d.getMonth()+1)).slice(-2)+'-'+('0'+d.getDate()).slice(-2)}
function tog(id){OPEN[id]=!OPEN[id];render()}

function quick(id,f,v){var b={};b[f]=v;
  api('/api/items/'+id,{method:'PUT',body:b}).then(function(){toast('저장했다');
    return Promise.all([meta(),load()])})}
function addResp(id){
  var c=$('rc'+id).value.trim();if(!c){toast('내용을 입력해라');return}
  if(!actor()){toast('상단에 작성자 이름부터 입력해라');return}
  api('/api/items/'+id+'/responses',{method:'POST',
    body:{responder:actor(),response_date:$('rd'+id).value,content:c}})
   .then(function(){toast('응답 등록');return load()})}
function editResp(rid){
  var cur='';for(var i=0;i<ITEMS.length;i++)for(var j=0;j<ITEMS[i].responses.length;j++)
    if(ITEMS[i].responses[j].id===rid)cur=ITEMS[i].responses[j].content;
  var v=prompt('응답 내용 수정',cur);if(v===null)return;
  api('/api/responses/'+rid,{method:'PUT',body:{content:v}}).then(function(){
    toast('수정했다');return load()})}
function delResp(rid){if(!confirm('이 응답 삭제할까?'))return;
  api('/api/responses/'+rid,{method:'DELETE'}).then(function(){return load()})}
function delItem(id){if(!confirm('요청 1건과 그 응답 전부 삭제된다. 진행할까?'))return;
  api('/api/items/'+id,{method:'DELETE'}).then(function(){toast('삭제했다');
    return Promise.all([meta(),load()])})}
function showHist(id){
  var el=$('h'+id);
  if(el.style.display!=='none'){el.style.display='none';return}
  el.style.display='block';el.innerHTML='불러오는 중...';
  api('/api/items/'+id+'/history').then(function(d){
    if(!d.history.length){el.innerHTML='이력 없음';return}
    el.innerHTML='<b>수정 이력</b>'+d.history.map(function(h){
      var txt=h.ts+' · '+(h.actor||'-')+' · '+h.action+
        (h.field?' ['+esc(h.field)+']':'')+
        (h.old_value?' : '+esc(h.old_value)+' → '+esc(h.new_value||''):
         (h.new_value?' : '+esc(h.new_value):''));
      return '<div>'+txt+'</div>'}).join('')})}

function openItem(id){
  var it=null;for(var i=0;i<ITEMS.length;i++)if(ITEMS[i].id===id)it=ITEMS[i];
  var v=it||{category:'요청',requester:actor(),request_date:today(),content:'',
             target:'',status:'대기',applied_date:''};
  $('modal').innerHTML='<div class="modal" onclick="if(event.target===this)closeM()">'+
   '<div class="box"><h2>'+(it?'요청 수정 (No.'+it.seq+')':'새 요청 등록')+'</h2>'+
   '<div class="g2"><div class="f"><label>구분</label><select id="m_cat">'+
     CATS.map(function(c){return '<option'+(c===v.category?' selected':'')+'>'+esc(c)+'</option>'}).join('')+
   '</select></div><div class="f"><label>상태</label><select id="m_st">'+
     STS.map(function(c){return '<option'+(c===v.status?' selected':'')+'>'+esc(c)+'</option>'}).join('')+
   '</select></div></div>'+
   '<div class="g2"><div class="f"><label>요청자</label>'+
     '<input id="m_req" value="'+esc(v.requester)+'" placeholder="예: 김윤환TL님"></div>'+
   '<div class="f"><label>요청일</label><input type="date" id="m_date" value="'+
     esc(v.request_date||today())+'"></div></div>'+
   '<div class="f"><label>대상 시스템 / FAB 태그 (콤마 구분)</label>'+
     '<input id="m_tag" list="taglist" value="'+esc(v.target)+
     '" placeholder="예: M16HUB, R-A룰, FAB위험도"></div>'+
   '<datalist id="taglist">'+TAGS.map(function(t){return '<option value="'+esc(t)+'">'}).join('')+'</datalist>'+
   '<div class="f"><label>내용</label><textarea id="m_ct">'+esc(v.content)+'</textarea></div>'+
   '<div class="f" style="width:50%"><label>적용일</label><input type="date" id="m_ad" value="'+
     esc(v.applied_date||'')+'"></div>'+
   '<div style="text-align:right;margin-top:6px"><button onclick="closeM()">취소</button> '+
   '<button class="p" onclick="saveItem('+(it?it.id:0)+')">저장</button></div>'+
   '</div></div>'}
function closeM(){$('modal').innerHTML=''}
function saveItem(id){
  var b={category:$('m_cat').value,status:$('m_st').value,requester:$('m_req').value,
         request_date:$('m_date').value,target:$('m_tag').value,content:$('m_ct').value,
         applied_date:$('m_ad').value};
  if(!b.content.trim()){toast('내용을 입력해라');return}
  var p=id?api('/api/items/'+id,{method:'PUT',body:b})
          :api('/api/items',{method:'POST',body:b});
  p.then(function(){closeM();toast('저장했다');return Promise.all([meta(),load()])})}

function doExport(){window.location='/api/export.xlsx?'+qs()}
function doImport(el){
  if(!el.files.length)return;
  if(!confirm('['+el.files[0].name+'] 를 가져온다.\n기존 데이터에 "추가"되고, 덮어쓰지 않는다.\n같은 파일을 두 번 올리면 중복 등록되니 주의해라.\n\n진행할까?')){el.value='';return}
  var fd=new FormData();fd.append('file',el.files[0]);
  fetch('/api/import',{method:'POST',body:fd,headers:{'X-Actor':encodeHeader(actor())}})
   .then(function(r){return r.json()}).then(function(d){
     el.value='';
     if(!d.ok){toast('실패: '+d.error);return}
     toast('가져오기 완료 — 요청 '+d.items+'건 / 응답 '+d.responses+'건');
     if(d.note)setTimeout(function(){toast(d.note)},2300);
     return Promise.all([meta(),load()])})}

$('actor').value=localStorage.getItem('reqlog_actor')||'';
$('actor').addEventListener('change',function(){
  localStorage.setItem('reqlog_actor',$('actor').value.trim())});
document.addEventListener('keydown',function(e){
  if(e.key!=='Escape')return;
  var lb=document.querySelector('.lightbox');
  if(lb){lb.remove();return}
  closeM()});
meta().then(load);
</script>
</body>
</html>
"""


if __name__ == "__main__":
    init_db()
    print("=" * 58)
    print(" 요청/응답 관리 시스템")
    print(" DB   : %s" % DB_PATH)
    print(" URL  : http://%s:%d  (내부망 접속용)" % (HOST, PORT))
    print("=" * 58)
    sys.stdout.flush()
    app.run(host=HOST, port=PORT, debug=False, threaded=True)
